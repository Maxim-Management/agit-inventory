"""
Импорт начальных данных из исходного файла Excel заказчика
("1. AGIT Spares and Components.xlsx") в БД сервиса.

Источники:
  - лист "AGIT Spares"      -> справочник деталей (parts) + история поступлений
                                (по колонкам "Received on ... as per order N"),
                                а для несерийных (расходных) деталей — остаток
                                "на начало" (opening_balance_qty = сумма Q-ty
                                по всем строкам детали) и цена за единицу в
                                рублях (unit_transfer_price_rub, среднее по
                                строкам детали). Для каждой партии/строки
                                дополнительно сохраняются справочные поля:
                                дата производства, таможенная пошлина (%),
                                курс валюты на момент покупки, общая
                                стоимость партии в CNY (receipts.date_mfg /
                                customs_duty_percent / exchange_rate /
                                total_cost_cny). Серийный номер по строке
                                партии из исходника (колонка r[3] листа
                                "AGIT Spares") в receipts НЕ сохраняется
                                (поле receipts.batch_serial_number удалено,
                                см. миграцию 0012) — используется только
                                транзитно, в памяти скрипта, чтобы если он
                                совпадает с реальной серийной единицей из
                                "AGIT components", связать с ней поступление
                                (receipts.unit_id).
  - лист "AGIT components"  -> серийные единицы (units), включая OD/ID для
                                ротора/статора (парсится из колонки Remarks),
                                статус (на складе / установлен на сборку); из
                                колонки "Installed on tool" дополнительно
                                создаётся реестр инструментов (tools) — по
                                одной записи на уникальный серийный номер
                                инструмента, с проставлением units.installed_on_tool_id

Деталь считается серийной (is_serialized), если для неё есть хотя бы одна
запись в листе "AGIT components" — именно там ведётся серийный учёт по
факту в исходном файле заказчика. Все остальные детали (расходники: масло,
уплотнения, кольца и т.п.) учитываются по количеству — так же, как серийные,
только без индивидуальных карточек, через единый механизм остатков
(см. app/stock.py) и списаний по количеству.

Примечание: исходный лист "AGIT Spares write-off" описывает не брак/ремонт/
повреждение, а расход деталей на сборку инструментов (ASSY of NNNNNN) —
компонент при этом остаётся В ЭКСПЛУАТАЦИИ (статус 'installed', установлен
на инструмент), а не списан со склада безвозвратно. Это уже отражено через
поле "Установлен на" / installed_on_tool_id у соответствующих серийных
единиц листа "AGIT components". Если бы этот же расход ещё и продублировать
в write_offs (статус 'written_off'), это противоречило бы факту, что деталь
реально работает на инструменте — поэтому отдельно не дублируется. Таблица
списаний (с причинами ремонт/брак/повреждение — то есть окончательное
выбытие детали) в системе стартует пустой и наполняется пользователями по
факту реальных списаний в процессе эксплуатации.

Скрипт идемпотентен: повторный запуск не создаёт дублей деталей/единиц.

Запуск:
  python scripts/import_excel.py /path/to/1._AGIT_Spares_and_Components.xlsx
"""
import os
import re
import sys
from datetime import date

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from dotenv import load_dotenv
load_dotenv()

import openpyxl  # noqa: E402
from app import db  # noqa: E402
from app.stock import STOCK_BASELINE_DATE  # noqa: E402

RECEIPT_COLUMNS = [
    # (индекс колонки 0-based, дата поступления, order_ref)
    (14, date(2026, 2, 1), "order 1"),
    (15, date(2026, 6, 1), "order 2"),
    (19, date(2026, 8, 1), "order 3"),
    (20, date(2026, 8, 1), "order 4"),
]

OD_RE = re.compile(r"OD\s*=\s*([\d,.]+)", re.IGNORECASE)
ID_RE = re.compile(r"\bID\s*=\s*([\d,.]+)", re.IGNORECASE)

assert max(d for _, d, _ in RECEIPT_COLUMNS) == STOCK_BASELINE_DATE, (
    "STOCK_BASELINE_DATE (app/stock.py) должна ТОЧНО совпадать с самой поздней "
    "исторической датой поступления ниже — иначе остаток несерийных деталей "
    "либо задвоится, либо (что хуже) новые операции будут молча игнорироваться."
)


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def norm_part_number(v):
    if v is None:
        return None
    return str(v).strip()


def guess_category(part_name):
    """Категория 'rotor'/'stator' присваивается только точному совпадению
    названия детали (без 'Rotor end stop' и т.п. похожих деталей)."""
    n = (part_name or "").strip().lower()
    if n == "rotor":
        return "rotor"
    if n == "stator":
        return "stator"
    return "other"


def parse_od_id(remarks):
    if not remarks:
        return None, None
    s = str(remarks)
    od_m = OD_RE.search(s)
    id_m = ID_RE.search(s)
    od = to_float(od_m.group(1)) if od_m else None
    idm = to_float(id_m.group(1)) if id_m else None
    return od, idm


def import_parts(ws):
    """Лист 'AGIT Spares' -> справочник деталей + история поступлений.
    Возвращает (part_id_by_number, receipt_link_candidates) — второй элемент
    список (receipt_id, part_id, serial_for_linking) для последующей
    backfill_receipt_unit_links() (серийный номер строки партии из Excel в
    БД не хранится, см. комментарий в шапке файла)."""
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    # Первый проход: агрегируем по part_number (в исходнике одна деталь может
    # занимать несколько строк — разные партии/цены/даты).
    agg = {}
    for r in rows:
        part_name = r[1]
        if part_name is None:
            continue
        part_number = norm_part_number(r[2])
        if not part_number or part_number == "#N/A":
            continue

        a = agg.setdefault(part_number, {
            "part_name": str(part_name).strip(),
            "tool_size": str(r[0]).strip() if r[0] else "",
            "specification": str(r[4]).strip() if r[4] else "",
            "standard_cost_list": [],
            "transfer_price_list": [],
            "unit_weight_list": [],
            "exchange_rate_list": [],
            "duty_percent_list": [],
            "qty_sum": 0.0,
            "rows": [],
        })
        std_cost = to_float(r[6])
        transfer_price = to_float(r[7])
        unit_weight = to_float(r[11])
        qty = to_float(r[13]) or 0.0
        row_tax_fraction = to_float(r[10]) if len(r) > 10 else None
        row_exchange_rate_agg = to_float(r[12]) if len(r) > 12 else None
        if std_cost is not None:
            a["standard_cost_list"].append(std_cost)
        if transfer_price is not None:
            a["transfer_price_list"].append(transfer_price)
        if unit_weight is not None:
            a["unit_weight_list"].append(unit_weight)
        if row_exchange_rate_agg is not None:
            a["exchange_rate_list"].append(row_exchange_rate_agg)
        if row_tax_fraction is not None:
            a["duty_percent_list"].append(round(row_tax_fraction * 100, 3))
        a["qty_sum"] += qty
        a["rows"].append(r)

    created, updated = 0, 0
    part_id_by_number = {}
    receipt_link_candidates = []

    for part_number, a in agg.items():
        category = guess_category(a["part_name"])
        min_stock = 1 if category in ("rotor", "stator") else 0
        standard_cost = sum(a["standard_cost_list"]) / len(a["standard_cost_list"]) if a["standard_cost_list"] else None
        transfer_price = sum(a["transfer_price_list"]) / len(a["transfer_price_list"]) if a["transfer_price_list"] else None
        unit_weight = sum(a["unit_weight_list"]) / len(a["unit_weight_list"]) if a["unit_weight_list"] else None
        exchange_rate = sum(a["exchange_rate_list"]) / len(a["exchange_rate_list"]) if a["exchange_rate_list"] else None
        duty_percent = sum(a["duty_percent_list"]) / len(a["duty_percent_list"]) if a["duty_percent_list"] else None

        existing = db.query_one("SELECT id FROM parts WHERE part_number = %s", [part_number])
        if existing:
            part_id = existing["id"]
            updated += 1
        else:
            part_id = db.execute_returning_id(
                """INSERT INTO parts (tool_size, part_name, part_number, category, specification,
                                       unit_weight_kg, standard_cost_cny, exchange_rate, customs_duty_percent,
                                       unit_transfer_price_rub, min_stock_qty, is_serialized, opening_balance_qty)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s) RETURNING id""",
                [a["tool_size"], a["part_name"], part_number, category, a["specification"],
                 unit_weight, standard_cost, exchange_rate, duty_percent, transfer_price, min_stock, a["qty_sum"]],
            )
            created += 1
        part_id_by_number[part_number] = part_id

        for r in a["rows"]:
            serial_raw = r[3] if len(r) > 3 else None
            serial_for_linking = str(serial_raw).strip() if serial_raw is not None else "-"
            row_date_mfg = str(r[5]).strip() if len(r) > 5 and r[5] else ""
            # Таможенная пошлина здесь НЕ пишется на receipts — она атрибут
            # позиции (parts.customs_duty_percent, усреднена по всем строкам
            # выше, в duty_percent) и одна для всех партий этой детали.
            row_exchange_rate = to_float(r[12]) if len(r) > 12 else None
            row_total_cost_cny = to_float(r[9]) if len(r) > 9 else None
            row_transfer_price = to_float(r[7]) if len(r) > 7 else None

            for col_idx, receipt_date, order_ref in RECEIPT_COLUMNS:
                if col_idx >= len(r):
                    continue
                qty = to_float(r[col_idx])
                if qty and qty > 0:
                    # remaining_quantity стартует равным quantity — при импорте
                    # с чистой БД ничего из этой партии ещё не списано; для
                    # серийных единиц backfill_receipt_unit_links() ниже
                    # уточнит его по фактическому статусу конкретной единицы.
                    receipt_id = db.execute_returning_id(
                        """INSERT INTO receipts (part_id, quantity, remaining_quantity, receipt_date, order_ref,
                                                  date_mfg, exchange_rate, total_cost_cny, transfer_price_rub, note)
                           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                        [part_id, qty, qty, receipt_date.isoformat(), order_ref,
                         row_date_mfg, row_exchange_rate, row_total_cost_cny, row_transfer_price,
                         "Импортировано из Excel"],
                    )
                    if serial_for_linking and serial_for_linking != "-":
                        receipt_link_candidates.append((receipt_id, part_id, serial_for_linking))

    print(f"Справочник деталей: создано {created}, уже существовало {updated}.")
    return part_id_by_number, receipt_link_candidates


def get_or_create_tool(serial_number, tool_size):
    """Инструмент (буровая компоновка) с данным серийным номером — создаёт
    запись в tools при первом упоминании в листе 'AGIT components' (колонка
    'Installed on tool'), иначе возвращает id существующей."""
    existing = db.query_one("SELECT id FROM tools WHERE serial_number = %s", [serial_number])
    if existing:
        return existing["id"]
    return db.execute_returning_id(
        """INSERT INTO tools (serial_number, tool_size, status, note)
           VALUES (%s,%s,'active',%s) RETURNING id""",
        [serial_number, tool_size, "Создан автоматически при импорте из Excel"],
    )


def import_units(ws, part_id_by_number):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created, skipped = 0, 0
    tool_id_by_serial = {}

    for r in rows:
        part_name = r[1]
        if part_name is None:
            continue
        tool_size = str(r[0]).strip() if r[0] else ""
        part_number = norm_part_number(r[2])
        serial_raw = r[3]
        serial_number = str(serial_raw).strip() if serial_raw is not None else ""
        specification = str(r[4]).strip() if r[4] else ""
        date_mfg = str(r[5]).strip() if r[5] else ""
        installed_on_raw = r[6]
        remarks = r[9]

        if not part_number or part_number == "#N/A" or not serial_number or serial_number == "-":
            skipped += 1
            continue

        part_id = part_id_by_number.get(part_number)
        if part_id is None:
            category = guess_category(part_name)
            existing = db.query_one("SELECT id FROM parts WHERE part_number = %s", [part_number])
            if existing:
                part_id = existing["id"]
            else:
                part_id = db.execute_returning_id(
                    """INSERT INTO parts (tool_size, part_name, part_number, category, specification, min_stock_qty, is_serialized)
                       VALUES (%s,%s,%s,%s,%s,%s,TRUE) RETURNING id""",
                    [tool_size, str(part_name).strip(), part_number, category, specification,
                     1 if category in ("rotor", "stator") else 0],
                )
            part_id_by_number[part_number] = part_id

        installed_on = "" if (installed_on_raw is None or str(installed_on_raw).strip().upper() == "NO") \
            else str(installed_on_raw).strip()
        status = "installed" if installed_on else "in_stock"
        od_mm, id_mm = parse_od_id(remarks)

        installed_on_tool_id = None
        if installed_on:
            if installed_on not in tool_id_by_serial:
                tool_id_by_serial[installed_on] = get_or_create_tool(installed_on, tool_size)
            installed_on_tool_id = tool_id_by_serial[installed_on]

        existing_unit = db.query_one(
            "SELECT id FROM units WHERE part_id = %s AND serial_number = %s", [part_id, serial_number]
        )
        if existing_unit:
            skipped += 1
            continue

        try:
            db.execute(
                """INSERT INTO units (part_id, serial_number, date_mfg, status, installed_on,
                                       installed_on_tool_id, od_mm, id_mm, remarks)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                [part_id, serial_number, date_mfg, status, installed_on, installed_on_tool_id, od_mm, id_mm,
                 str(remarks).strip() if remarks else ""],
            )
            created += 1
        except Exception as e:
            print(f"  Пропущена единица {part_number}/{serial_number}: {e}")
            skipped += 1

    print(f"Серийные компоненты: создано {created}, пропущено {skipped}.")
    print(f"Инструменты: создано записей в реестре {len(tool_id_by_serial)} "
          f"(уникальных серийных номеров 'Installed on tool' в этом листе).")


def backfill_receipt_unit_links(receipt_link_candidates):
    """Проставляет receipts.unit_id (и, в обратную сторону, units.receipt_id —
    основную связь для партионного учёта, см. миграцию 0008) там, где серийный
    номер строки поступления (из листа 'AGIT Spares', собран в памяти в
    import_parts() — сама БД это поле больше не хранит, см. миграцию 0012)
    совпадает с реальной серийной единицей той же детали, заведённой из листа
    'AGIT components'. Заодно уточняет остаток партии (remaining_quantity):
    для партии из одной привязанной единицы остаток = 1, если единица ещё
    'in_stock', иначе 0 (уже установлена/списана — в партии ничего не осталось)."""
    linked = 0
    for receipt_id, part_id, serial in receipt_link_candidates:
        unit = db.query_one(
            "SELECT id, status FROM units WHERE part_id = %s AND serial_number = %s",
            [part_id, serial],
        )
        if unit:
            remaining = 1 if unit["status"] == "in_stock" else 0
            db.execute(
                "UPDATE receipts SET unit_id = %s, remaining_quantity = %s WHERE id = %s",
                [unit["id"], remaining, receipt_id],
            )
            db.execute("UPDATE units SET receipt_id = %s WHERE id = %s", [receipt_id, unit["id"]])
            linked += 1
    print(f"Поступления связаны с серийными единицами: {linked} из {len(receipt_link_candidates)} с указанным серийным номером.")


def finalize_is_serialized():
    """Деталь серийная, только если для неё реально есть карточки в units —
    так поступивший из 'AGIT Spares' расходник (масло, уплотнения) без единиц
    в 'AGIT components' автоматически становится учитываемым по количеству."""
    serialized_ids = [r["part_id"] for r in db.query_all("SELECT DISTINCT part_id FROM units")]
    all_ids = [r["id"] for r in db.query_all("SELECT id FROM parts")]
    non_serialized_ids = [i for i in all_ids if i not in set(serialized_ids)]

    for pid in serialized_ids:
        db.execute("UPDATE parts SET is_serialized = %s WHERE id = %s", [True, pid])
    for pid in non_serialized_ids:
        db.execute("UPDATE parts SET is_serialized = %s WHERE id = %s", [False, pid])

    print(f"Способ учёта: серийных деталей {len(serialized_ids)}, по количеству {len(non_serialized_ids)}.")


def main():
    if len(sys.argv) < 2:
        print("Использование: python scripts/import_excel.py /path/to/file.xlsx")
        sys.exit(1)
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, data_only=True)

    # Одно соединение на весь импорт вместо сотен отдельных — быстрее и
    # устойчивее к обрыву соединения пулером Supabase при частом открытии
    # новых подключений подряд (см. db.bulk_session).
    with db.bulk_session():
        part_id_by_number, receipt_link_candidates = import_parts(wb["AGIT Spares"])
        import_units(wb["AGIT components"], part_id_by_number)
        backfill_receipt_unit_links(receipt_link_candidates)
        finalize_is_serialized()

    print("Импорт завершён.")


if __name__ == "__main__":
    main()
